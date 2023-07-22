import time
from classes.reports import newReport
import logging
import openpyxl

logging.basicConfig(
    level=logging.WARNING,  # Set the log level (DEBUG, INFO, WARNING, ERROR, CRITICAL)
    format='%(asctime)s - %(levelname)s - %(message)s',  # Set the log message format
    filename='app reference/app.log',  # Specify the log file name
    filemode='a'  # Set the file mode (w: write, a: append)
)

file_path = "app reference/reports.xlsx"
report_sheet_name = "reports"
mp_sheet_name = "marketplaces"

workbook = openpyxl.load_workbook(file_path)
report_sheet = workbook[report_sheet_name]
mp_sheet = workbook[mp_sheet_name]

reports = []
for row in report_sheet.iter_rows(min_row=1, values_only=True):
    cell_value = row[0]
    if cell_value is not None:
        reports.append(cell_value)

markets = []
for row in mp_sheet.iter_rows(min_row=1, values_only=True):
    cell_value = row[0]
    if cell_value is not None:
        markets.append(cell_value)
workbook.close()

def run_auto(reports, markets):
    # Create a list to store the objects
    report_objects = []
    excluded_report_objects = []

    for market in markets:
        for report in reports:
            new_report = newReport(f"{report}", f"{market}")
            if new_report.request_report() == 202:
                report_objects.append(new_report)
            else:
                excluded_report_objects.append(new_report)
            time.sleep(2)

    done_processing_report_objects = []
    no_of_reports = len(report_objects)

    print(report_objects)
    print(excluded_report_objects)

    while no_of_reports != len(done_processing_report_objects):
        if len(report_objects) >= 10:
            sleep_sec = 3
        elif len(report_objects) >= 5:
            sleep_sec = 7
        else:
            sleep_sec = 15
        for report in report_objects:
            report.get_report(report.report_id)

            if report.report_status == 'FATAL':
                report_objects.remove(report)
                no_of_reports -= 1

                error_message = f"Removed {report.report_name} from downloads, report Status: {report.report_status}"
                logging.error(error_message)
                print(error_message)
                # raise Exception(error_message)

            elif report.report_status == 'DONE':
                done_processing_report_objects.append(report)
                report_objects.remove(report)
                if report.doc_id != "":
                    report.get_report_link(report.doc_id)
                    report.download_report()
                    report.parse_and_update_report()
                else:
                    print("DONE but no doc_id???")
            else:
                pass
            time.sleep(sleep_sec)


run_auto(reports, markets)


